# -*- coding: utf-8 -*-
"""
Atualiza calcinhas a partir de catalogo_caixa_secreta_novo.xlsx + pasta CALCINHA.

Campos: Meu Preço Venda, Foto/Foto Verso (match por nome), Tamanho, Cores (Sabores).

Uso:
  python scripts/importar_calcinhas_novo.py
  python scripts/importar_calcinhas_novo.py --sem-processar-imagem
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from PIL import Image

from imagem_fundo import save_transparent_catalog_png

BASE = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"C:\Users\Ikaro\post\catalogo_caixa_secreta.xlsx")
DEFAULT_IMAGES = Path(r"C:\Users\Ikaro\post\imagens_caixa_secreta\CALCINHA")
LINGerie_JSON = BASE / "frontend" / "src" / "data" / "catalogo-lingerie.json"
IMG_OUT = BASE / "frontend" / "public" / "importados" / "catalogo" / "calcinhas"
LOG_FILE = BASE / "dados" / "import_calcinhas.log"
SHEET = "CALCINHA"
SLUG = "calcinhas"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def norm_key(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def norm_ref(ref) -> str:
    if ref is None:
        return ""
    return str(ref).strip().upper()


def slugify(text: str, max_len: int = 48) -> str:
    s = unicodedata.normalize("NFKD", text or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return (s or "item")[:max_len]


def parse_price(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def header_index(header: tuple) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        key = norm_key(str(h))
        if "produto" in key and "nome" not in out:
            out["nome"] = i
        elif key == "foto" and "verso" not in str(h).lower():
            out["foto"] = i
        elif "verso" in key or "fotoverso" in key:
            out["foto_verso"] = i
        elif key == "ref":
            out["ref"] = i
        elif "tamanho" in key:
            out["tamanho"] = i
        elif "sabor" in key or key == "cores" or ("cor" in key and "categoria" not in key):
            out["cores"] = i
        elif "preco" in key and "venda" in key:
            out["preco"] = i
        elif "promocao" in key or ("3" in key and "por" in key):
            out["combo3"] = i
    return out


def cell(row: tuple, idx: dict, key: str):
    if key not in idx or idx[key] >= len(row):
        return None
    v = row[idx[key]]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_list_field(text: str | None, prefix: str) -> list[str]:
    if not text:
        return []
    s = str(text).strip()
    s = re.sub(rf"^{prefix}\s*:?\s*", "", s, flags=re.I)
    parts = re.split(r"[,;/]", s)
    return [p.strip() for p in parts if p.strip()]


def build_description(tamanho: str | None, cores: str | None) -> str:
    parts = []
    if tamanho:
        parts.append(tamanho.strip())
    if cores:
        parts.append(cores.strip())
    return "\n".join(parts)


def strip_side_label(stem: str) -> str:
    return re.sub(r"\s*-\s*(frente|verso)\s*$", "", stem, flags=re.I).strip()


def name_score(product_name: str, file_stem: str) -> float:
    prod = norm_key(product_name)
    stem = norm_key(strip_side_label(file_stem))
    if not prod or not stem:
        return 0.0
    if prod == stem or prod in stem or stem in prod:
        return 0.98
    tokens = [t for t in re.split(r"[^a-z0-9]+", product_name.lower()) if len(t) >= 4]
    if not tokens:
        return 0.0
    hits = sum(1 for t in tokens if t in stem)
    return hits / len(tokens)


def list_calcinha_images(images_root: Path) -> tuple[list[Path], list[Path]]:
    frente, verso = [], []
    for p in images_root.iterdir():
        if not p.is_file() or p.suffix.lower() not in IMAGE_EXTS:
            continue
        low = p.stem.lower()
        if "verso" in low:
            verso.append(p)
        elif "frente" in low:
            frente.append(p)
        else:
            frente.append(p)
    return frente, verso


def find_side(
    product_name: str,
    ref: str | None,
    foto_hint: str | None,
    pool: list[Path],
    *,
    side: str,
) -> Path | None:
    if foto_hint:
        hint = Path(foto_hint).name.lower()
        for p in pool:
            if hint in p.name.lower():
                return p

    ref_n = norm_ref(ref)
    if ref_n:
        for p in pool:
            if ref_n in p.stem.upper():
                return p

    scored = [(name_score(product_name, p.stem), p) for p in pool]
    scored.sort(key=lambda x: -x[0])
    if scored and scored[0][0] >= 0.55:
        return scored[0][1]

    # fallback: único arquivo com side no nome próximo ao produto
    side_key = side.lower()
    for score, p in scored:
        if score >= 0.35 and side_key in p.stem.lower():
            return p
    return None


def process_image(src: Path, dest: Path, *, raw_copy: bool) -> bool:
    try:
        im = Image.open(src)
    except OSError:
        return False
    if raw_copy:
        dest.parent.mkdir(parents=True, exist_ok=True)
        if src.suffix.lower() == ".png":
            shutil.copy2(src, dest)
            return dest.exists()
        im.convert("RGB").save(dest.with_suffix(".png"), format="PNG", optimize=True)
        return dest.with_suffix(".png").exists()
    return save_transparent_catalog_png(im, dest)


def read_rows(xlsx: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Aba {SHEET!r} não encontrada em {xlsx}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    idx = header_index(rows[0])
    if "nome" not in idx or "preco" not in idx:
        raise ValueError("Planilha CALCINHA: faltam colunas Produto e Meu Preço Venda")

    out = []
    for raw in rows[1:]:
        if not raw or all(c is None or str(c).strip() == "" for c in raw):
            continue
        nome = cell(raw, idx, "nome")
        if not nome:
            continue
        tamanho = cell(raw, idx, "tamanho")
        cores = cell(raw, idx, "cores")
        out.append(
            {
                "name": nome,
                "ref": cell(raw, idx, "ref"),
                "price": parse_price(raw[idx["preco"]] if "preco" in idx else 0),
                "combo3": parse_price(raw[idx["combo3"]]) if "combo3" in idx and raw[idx["combo3"]] else None,
                "tamanho": tamanho,
                "cores": cores,
                "sizes": parse_list_field(tamanho, "tam"),
                "colors": parse_list_field(cores, "cores"),
                "description": build_description(tamanho, cores),
                "foto": cell(raw, idx, "foto"),
                "foto_verso": cell(raw, idx, "foto_verso"),
            }
        )
    return out


def product_id(ref: str | None, name: str) -> str:
    base = slugify(norm_ref(ref) or name)
    return f"calcinhas-{base}"[:80]


def load_json(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: list[dict]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--imagens", type=Path, default=DEFAULT_IMAGES)
    ap.add_argument("--sem-processar-imagem", action="store_true", help="Copia JPG sem recorte")
    args = ap.parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(args.xlsx)
    if not args.imagens.exists():
        raise FileNotFoundError(args.imagens)

    sheet_rows = read_rows(args.xlsx)
    print(f"Linhas na planilha ({SHEET}): {len(sheet_rows)}")

    frente_files, verso_files = list_calcinha_images(args.imagens)
    print(f"Imagens: {len(frente_files)} frente, {len(verso_files)} verso")

    lingerie = load_json(LINGerie_JSON)
    outros = [p for p in lingerie if p.get("categorySlug") != SLUG]
    antigos = sum(1 for p in lingerie if p.get("categorySlug") == SLUG)
    print(f"Calcinhas antigas removidas do JSON: {antigos}")

    IMG_OUT.mkdir(parents=True, exist_ok=True)
    log: list[str] = []
    novos: list[dict] = []
    ok_img = 0

    for row in sheet_rows:
        ref = row.get("ref")
        name = row["name"]
        pid = product_id(ref, name)
        safe = slugify(norm_ref(ref) or name)[:50]

        src_frente = find_side(name, ref, row.get("foto"), frente_files, side="frente")
        src_verso = find_side(name, ref, row.get("foto_verso"), verso_files, side="verso")

        image_url = image_back = None
        if src_frente:
            dest = IMG_OUT / f"{safe}-frente.png"
            if process_image(src_frente, dest, raw_copy=args.sem_processar_imagem):
                image_url = f"/importados/catalogo/calcinhas/{dest.name}"
                ok_img += 1
            else:
                log.append(f"{name} | frente falhou: {src_frente.name}")
        else:
            log.append(f"{name} | ref={ref} | SEM FRENTE")

        if src_verso:
            dest_v = IMG_OUT / f"{safe}-verso.png"
            if process_image(src_verso, dest_v, raw_copy=args.sem_processar_imagem):
                image_back = f"/importados/catalogo/calcinhas/{dest_v.name}"
            else:
                log.append(f"{name} | verso falhou: {src_verso.name}")
        else:
            log.append(f"{name} | ref={ref} | SEM VERSO")

        now = utc_now()
        novos.append(
            {
                "id": pid,
                "name": name,
                "ref": ref,
                "price": row["price"],
                "category": "Calcinhas",
                "categorySlug": SLUG,
                "description": row["description"],
                "size": row.get("tamanho"),
                "colors": row.get("cores"),
                "sizes": row.get("sizes") or [],
                "colorsList": row.get("colors") or [],
                "image": image_url,
                "imageBack": image_back,
                "comboThree": row.get("combo3"),
                "page": None,
                "source": "catalogo_caixa_secreta_novo.xlsx — CALCINHA",
                "importSource": "planilha-nova",
                "published": bool(image_url),
                "active": True,
                "createdAt": now,
                "updatedAt": now,
            }
        )

    final = outros + novos
    save_json(LINGerie_JSON, final)

    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    LOG_FILE.write_text("\n".join(log) if log else "(ok)\n", encoding="utf-8")

    pub = sum(1 for p in novos if p.get("published"))
    print(f"Novas calcinhas: {len(novos)} | Publicadas (com frente): {pub}")
    print(f"JSON lingerie: {len(final)} itens ({len(outros)} outras categorias)")
    print(f"Log: {LOG_FILE}")


if __name__ == "__main__":
    main()
