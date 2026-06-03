# -*- coding: utf-8 -*-
"""
Legado: use importar_catalogo_caixa_secreta.py (importa todo o catálogo).
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from PIL import Image

from imagem_fundo import save_transparent_catalog_png

BASE = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"C:\Users\Ikaro\post\catalogo_caixa_secreta.xlsx")
DEFAULT_IMAGES = Path(r"C:\Users\Ikaro\post\imagens_caixa_secreta")
LINGerie_JSON = BASE / "frontend" / "src" / "data" / "catalogo-lingerie.json"
IMG_BASE = BASE / "frontend" / "public" / "importados" / "catalogo"
LOG_FILE = BASE / "dados" / "import_lingerie_planilha.log"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}

SHEET_CONFIG: dict[str, dict[str, str]] = {
    "CALCINHA": {
        "slug": "calcinhas",
        "title": "Calcinhas",
        "folder": "CALCINHA",
    },
    "LINGERIES": {
        "slug": "lingeries",
        "title": "Lingeries",
        "folder": "LINGERIES",
    },
    "ESPARTILHO": {
        "slug": "espartilhos",
        "title": "Espartilhos",
        "folder": "ESPARTILHO",
    },
}

IMPORT_SLUGS = frozenset(c["slug"] for c in SHEET_CONFIG.values())


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def norm_key(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def norm_ref(ref) -> str:
    if ref is None:
        return ""
    return str(ref).strip().upper()


def slugify(text: str, max_len: int = 48) -> str:
    s = unicodedata.normalize("NFKD", text or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return (s or "item")[:max_len]


def parse_price(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def header_index(header: tuple) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        key = norm_key(str(h))
        raw_h = str(h).lower()
        if "produto" in key and "nome" not in out:
            out["nome"] = i
        elif key == "foto" and "verso" not in raw_h:
            out["foto"] = i
        elif "verso" in key or "fotoverso" in key:
            out["foto_verso"] = i
        elif key == "ref":
            out["ref"] = i
        elif "tamanho" in key:
            out["tamanho"] = i
        elif "sabor" in key or key == "cores" or ("cor" in key and "categoria" not in key):
            out["cores"] = i
        elif "preco" in key and "venda" in key:
            out["preco"] = i
        elif "promocao" in key or ("3" in key and "por" in key):
            out["combo3"] = i
    return out


def cell(row: tuple, idx: dict, key: str):
    if key not in idx or idx[key] >= len(row):
        return None
    v = row[idx[key]]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_list_field(text: str | None, prefix: str) -> list[str]:
    if not text:
        return []
    s = str(text).strip()
    s = re.sub(rf"^{prefix}\s*:?\s*", "", s, flags=re.I)
    parts = re.split(r"[,;/]", s)
    return [p.strip() for p in parts if p.strip()]


def build_description(tamanho: str | None, cores: str | None) -> str:
    parts = []
    if tamanho:
        parts.append(tamanho.strip())
    if cores:
        parts.append(cores.strip())
    return "\n".join(parts)


def strip_side_label(stem: str) -> str:
    return re.sub(r"\s*-\s*(frente|verso)\s*$", "", stem, flags=re.I).strip()


def name_score(product_name: str, file_stem: str) -> float:
    prod = norm_key(product_name)
    stem = norm_key(strip_side_label(file_stem))
    if not prod or not stem:
        return 0.0
    if prod == stem or prod in stem or stem in prod:
        return 0.98
    tokens = [t for t in re.split(r"[^a-z0-9]+", product_name.lower()) if len(t) >= 4]
    if not tokens:
        return 0.0
    hits = sum(1 for t in tokens if t in stem)
    return hits / len(tokens)


def list_side_images(images_root: Path) -> tuple[list[Path], list[Path]]:
    frente, verso = [], []
    if not images_root.is_dir():
        return frente, verso
    for p in images_root.iterdir():
        if not p.is_file() or p.suffix.lower() not in IMAGE_EXTS:
            continue
        low = p.stem.lower()
        if "verso" in low:
            verso.append(p)
        elif "frente" in low:
            frente.append(p)
        else:
            frente.append(p)
    return frente, verso


def find_side(
    product_name: str,
    ref: str | None,
    foto_hint: str | None,
    pool: list[Path],
    *,
    side: str,
) -> Path | None:
    if foto_hint:
        hint = Path(foto_hint).name.lower()
        for p in pool:
            if hint in p.name.lower():
                return p

    ref_n = norm_ref(ref)
    if ref_n:
        for p in pool:
            if ref_n in p.stem.upper():
                return p

    scored = [(name_score(product_name, p.stem), p) for p in pool]
    scored.sort(key=lambda x: -x[0])
    if scored and scored[0][0] >= 0.55:
        return scored[0][1]

    side_key = side.lower()
    for score, p in scored:
        if score >= 0.35 and side_key in p.stem.lower():
            return p
    return None


def process_image(src: Path, dest: Path, *, raw_copy: bool) -> bool:
    try:
        im = Image.open(src)
    except OSError:
        return False
    if raw_copy:
        dest.parent.mkdir(parents=True, exist_ok=True)
        if src.suffix.lower() == ".png":
            shutil.copy2(src, dest)
            return dest.exists()
        im.convert("RGB").save(dest, format="PNG", optimize=True)
        return dest.exists()
    return save_transparent_catalog_png(im, dest)


def read_sheet(xlsx: Path, sheet_name: str) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Aba {sheet_name!r} não encontrada")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    idx = header_index(rows[0])
    if "nome" not in idx or "preco" not in idx:
        raise ValueError(f"{sheet_name}: faltam colunas Produto e Meu Preço Venda")

    out = []
    for raw in rows[1:]:
        if not raw or all(c is None or str(c).strip() == "" for c in raw):
            continue
        nome = cell(raw, idx, "nome")
        if not nome:
            continue
        tamanho = cell(raw, idx, "tamanho")
        cores = cell(raw, idx, "cores")
        out.append(
            {
                "name": nome,
                "ref": cell(raw, idx, "ref"),
                "price": parse_price(raw[idx["preco"]] if "preco" in idx else 0),
                "combo3": parse_price(raw[idx["combo3"]])
                if "combo3" in idx and raw[idx["combo3"]]
                else None,
                "tamanho": tamanho,
                "cores": cores,
                "sizes": parse_list_field(tamanho, "tam"),
                "colors": parse_list_field(cores, "cores"),
                "description": build_description(tamanho, cores),
                "foto": cell(raw, idx, "foto"),
                "foto_verso": cell(raw, idx, "foto_verso"),
            }
        )
    return out


def product_id(slug: str, ref: str | None, name: str) -> str:
    base = slugify(norm_ref(ref) or name)
    return f"{slug}-{base}"[:80]


def load_json(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: list[dict]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def import_sheet(
    *,
    sheet_name: str,
    cfg: dict[str, str],
    xlsx: Path,
    images_root: Path,
    raw_copy: bool,
    log: list[str],
) -> list[dict]:
    slug = cfg["slug"]
    title = cfg["title"]
    folder = images_root / cfg["folder"]
    img_out = IMG_BASE / slug
    img_out.mkdir(parents=True, exist_ok=True)

    rows = read_sheet(xlsx, sheet_name)
    frente_files, verso_files = list_side_images(folder)
    print(f"  {sheet_name}: {len(rows)} produtos | {len(frente_files)} frente, {len(verso_files)} verso")

    products = []
    for row in rows:
        ref = row.get("ref")
        name = row["name"]
        pid = product_id(slug, ref, name)
        safe = slugify(norm_ref(ref) or name)[:50]

        src_frente = find_side(name, ref, row.get("foto"), frente_files, side="frente")
        src_verso = find_side(name, ref, row.get("foto_verso"), verso_files, side="verso")

        image_url = image_back = None
        if src_frente:
            dest = img_out / f"{safe}-frente.png"
            if process_image(src_frente, dest, raw_copy=raw_copy):
                image_url = f"/importados/catalogo/{slug}/{dest.name}"
            else:
                log.append(f"[{sheet_name}] {name} | frente falhou")
        else:
            log.append(f"[{sheet_name}] {name} | ref={ref} | SEM FRENTE")

        if src_verso:
            dest_v = img_out / f"{safe}-verso.png"
            if process_image(src_verso, dest_v, raw_copy=raw_copy):
                image_back = f"/importados/catalogo/{slug}/{dest_v.name}"
            else:
                log.append(f"[{sheet_name}] {name} | verso falhou")
        elif src_frente:
            log.append(f"[{sheet_name}] {name} | ref={ref} | SEM VERSO")

        now = utc_now()
        products.append(
            {
                "id": pid,
                "name": name,
                "ref": ref,
                "price": row["price"],
                "category": title,
                "categorySlug": slug,
                "description": row["description"],
                "size": row.get("tamanho"),
                "colors": row.get("cores"),
                "sizes": row.get("sizes") or [],
                "colorsList": row.get("colors") or [],
                "image": image_url,
                "imageBack": image_back,
                "comboThree": row.get("combo3"),
                "page": None,
                "source": f"catalogo_caixa_secreta.xlsx — {sheet_name}",
                "importSource": "planilha-lingerie",
                "published": bool(image_url),
                "active": True,
                "createdAt": now,
                "updatedAt": now,
            }
        )
    pub = sum(1 for p in products if p.get("published"))
    print(f"  publicados: {pub}/{len(products)}")
    return products


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--imagens", type=Path, default=DEFAULT_IMAGES)
    ap.add_argument("--aba", help="Só uma aba: CALCINHA, LINGERIES ou ESPARTILHO")
    ap.add_argument("--sem-processar-imagem", action="store_true")
    args = ap.parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(args.xlsx)

    sheets = {args.aba: SHEET_CONFIG[args.aba]} if args.aba else SHEET_CONFIG
    if args.aba and args.aba not in SHEET_CONFIG:
        raise ValueError(f"Aba inválida. Use: {', '.join(SHEET_CONFIG)}")

    lingerie = load_json(LINGerie_JSON)
    if args.aba:
        slugs_remove = {SHEET_CONFIG[args.aba]["slug"]}
    else:
        slugs_remove = IMPORT_SLUGS
    outros = [p for p in lingerie if p.get("categorySlug") not in slugs_remove]
    removidos = len(lingerie) - len(outros)
    print(f"Itens antigos removidos ({', '.join(sorted(slugs_remove))}): {removidos}")

    log: list[str] = []
    novos: list[dict] = []
    for sheet_name, cfg in sheets.items():
        novos.extend(
            import_sheet(
                sheet_name=sheet_name,
                cfg=cfg,
                xlsx=args.xlsx,
                images_root=args.imagens,
                raw_copy=args.sem_processar_imagem,
                log=log,
            )
        )

    final = outros + novos
    save_json(LINGerie_JSON, final)

    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    LOG_FILE.write_text("\n".join(log) if log else "(ok)\n", encoding="utf-8")

    pub = sum(1 for p in novos if p.get("published"))
    print(f"Total importado: {len(novos)} | Com foto: {pub}")
    print(f"JSON lingerie: {len(final)} itens")
    print(f"Log: {LOG_FILE}")


if __name__ == "__main__":
    import subprocess
    import sys

    script = Path(__file__).resolve().parent / "importar_catalogo_caixa_secreta.py"
    raise SystemExit(subprocess.call([sys.executable, str(script), *sys.argv[1:]]))
