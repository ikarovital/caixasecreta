# -*- coding: utf-8 -*-
"""
Gera PDF no visual do site (cores + logo) com dados da planilha Excel.

- Textos e preços: sempre da planilha Catalogo_ClubeCaixaSecreta.xlsx
- Imagens: arquivos do site (catalogo JSON) por ref/nome; se não achar, foto embutida na planilha
"""
from __future__ import annotations

import argparse
import json
import re
import tempfile
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import fitz
import openpyxl

BASE = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "post" / "Catalogo_ClubeCaixaSecreta.xlsx"
DEFAULT_OUT = BASE / "dados" / "catalogo_clube_caixa_secreta.pdf"
LOGO = BASE / "imagens" / "logo.jpg"
FALLBACK_LOGO = BASE / "frontend" / "public" / "imagens" / "logo.jpg"
LINGerie_JSON = BASE / "frontend" / "src" / "data" / "catalogo-lingerie.json"
IMPORTADO_JSON = BASE / "frontend" / "src" / "data" / "catalogo-importado.json"
PUBLIC_ROOTS = (
    BASE / "frontend" / "public",
    BASE / "frontend" / "dist",
    BASE,
)
EXTRACAO_ROOT = BASE / "dados" / "extracao_pdf"

# Cores do site (frontend/tailwind.config.js)
BG = (13 / 255, 6 / 255, 24 / 255)
TEXT = (0.95, 0.91, 1.0)
TEXT_MUTED = (0.72, 0.68, 0.82)
ACCENT = (108 / 255, 43 / 255, 217 / 255)
CARD_BORDER = (108 / 255, 43 / 255, 217 / 255, 0.45)

SHEETS_ORDER = [
    "COMESTÍVEIS",
    "COSMÉTICOS",
    "VIBRADORES",
    "PRÓTESES",
    "VELAS",
]

COLS_PER_ROW = 2
ROWS_PER_PAGE = 2
PRODUCTS_PER_PAGE = COLS_PER_ROW * ROWS_PER_PAGE

MARGIN = 36
GAP = 14
DESC_MAX = 130
NAME_MAX = 52


@dataclass
class Product:
    name: str
    description: str
    price: str
    image_bytes: bytes | None
    ref: str
    image_source: str = ""


@dataclass
class SiteCatalogIndex:
    by_ref: dict[str, dict]
    by_name: dict[str, dict]


def norm_text(value) -> str:
    if value is None:
        return ""
    s = unicodedata.normalize("NFKD", str(value))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.strip().lower())


def norm_header(value) -> str:
    if value is None:
        return ""
    s = unicodedata.normalize("NFKD", str(value))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.strip().lower())


def brl(value) -> str:
    if value is None or value == "":
        return "Consulte"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return str(value)
    inteiro, cent = divmod(round(v * 100), 100)
    parte_int = f"{inteiro:,}".replace(",", ".")
    return f"R$ {parte_int},{cent:02d}"


def truncate(text: str, max_len: int) -> str:
    text = re.sub(r"\s+", " ", (text or "").strip())
    if len(text) <= max_len:
        return text
    return text[: max_len - 1].rstrip() + "…"


def sheet_key(name: str) -> str:
    s = norm_header(name)
    return re.sub(r"[^a-z0-9]", "", s)


def build_header_map(headers: tuple) -> dict[str, int]:
    return {norm_header(h): i for i, h in enumerate(headers) if h}


def cell(row: tuple, col_map: dict[str, int], *keys: str):
    for key in keys:
        idx = col_map.get(key)
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def short_description(row: tuple, col_map: dict[str, int]) -> str:
    desc = cell(row, col_map, "descricao", "descrição")
    if desc:
        return truncate(str(desc), DESC_MAX)

    parts: list[str] = []
    ref = cell(row, col_map, "ref")
    tam = cell(row, col_map, "tamanho")
    sab = cell(row, col_map, "sabores")
    ali = cell(row, col_map, "alimentacao", "alimentação")

    if ref:
        parts.append(f"Ref. {ref}")
    if tam:
        parts.append(str(tam))
    if sab and str(sab).strip().lower() not in ("único", "unico", ""):
        parts.append(str(sab))
    if ali:
        parts.append(str(ali))

    return truncate(" · ".join(parts), DESC_MAX) if parts else ""


def load_site_catalog_index() -> SiteCatalogIndex:
    by_ref: dict[str, dict] = {}
    by_name: dict[str, dict] = {}
    for path in (LINGerie_JSON, IMPORTADO_JSON):
        if not path.exists():
            continue
        items = json.loads(path.read_text(encoding="utf-8"))
        for product in items:
            ref = str(product.get("ref") or "").strip().upper()
            name = norm_text(product.get("name"))
            if ref and ref not in by_ref:
                by_ref[ref] = product
            if name and name not in by_name:
                by_name[name] = product
    return SiteCatalogIndex(by_ref=by_ref, by_name=by_name)


def resolve_public_path(url_path: str) -> Path | None:
    if not url_path:
        return None
    rel = url_path.lstrip("/").replace("/", "\\")
    if rel.startswith("extracao_pdf"):
        candidate = EXTRACAO_ROOT / rel[len("extracao_pdf") :].lstrip("\\/")
        if candidate.exists():
            return candidate
    for root in PUBLIC_ROOTS:
        candidate = root / rel.replace("\\", "/")
        if candidate.is_file():
            return candidate
    return None


def read_file_bytes(path: Path) -> bytes | None:
    try:
        return path.read_bytes()
    except OSError:
        return None


def site_product_image(site: SiteCatalogIndex, ref: str, name: str) -> tuple[bytes | None, str]:
    product = None
    ref_key = ref.strip().upper()
    if ref_key:
        product = site.by_ref.get(ref_key)
    if product is None:
        product = site.by_name.get(norm_text(name))
    if not product:
        return None, ""
    image_url = product.get("image")
    if not image_url:
        return None, ""
    path = resolve_public_path(str(image_url))
    if not path:
        return None, ""
    data = read_file_bytes(path)
    return data, f"site:{path.name}" if data else (None, "")


def pick_image(
    site: SiteCatalogIndex,
    ref: str,
    name: str,
    excel_bytes: bytes | None,
) -> tuple[bytes | None, str]:
    site_bytes, site_label = site_product_image(site, ref, name)
    if site_bytes:
        return site_bytes, site_label
    if excel_bytes:
        return excel_bytes, "planilha"
    return None, ""


def extract_row_images(ws) -> dict[int, bytes]:
    out: dict[int, bytes] = {}
    for img in getattr(ws, "_images", []):
        anchor = getattr(img, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue
        excel_row = int(marker.row) + 1
        if excel_row not in out:
            try:
                out[excel_row] = img._data()
            except Exception:
                pass
    return out


def read_products(
    ws_values,
    ws_images,
    site: SiteCatalogIndex,
) -> list[Product]:
    rows = list(ws_values.iter_rows(values_only=True))
    if not rows:
        return []

    col_map = build_header_map(rows[0])
    images = extract_row_images(ws_images) if ws_images else {}
    products: list[Product] = []

    for excel_row, row in enumerate(rows[1:], start=2):
        name = cell(row, col_map, "produto")
        if not name:
            continue
        price_val = cell(row, col_map, "meu preco venda", "meu preço venda")
        ref = str(cell(row, col_map, "ref") or "")
        image_bytes, image_source = pick_image(
            site, ref, str(name), images.get(excel_row)
        )
        products.append(
            Product(
                name=truncate(str(name), NAME_MAX),
                description=short_description(row, col_map),
                price=brl(price_val),
                image_bytes=image_bytes,
                ref=ref,
                image_source=image_source,
            )
        )
    return products


def logo_path() -> Path | None:
    for p in (LOGO, FALLBACK_LOGO):
        if p.exists():
            return p
    return None


def fill_background(page: fitz.Page) -> None:
    page.draw_rect(page.rect, color=BG, fill=BG)


def draw_category_page(doc: fitz.Document, title: str) -> None:
    page = doc.new_page(width=595, height=842)
    fill_background(page)
    r = page.rect
    lp = logo_path()
    if lp:
        rect = fitz.Rect(r.width / 2 - 48, 120, r.width / 2 + 48, 216)
        page.insert_image(rect, filename=str(lp))

    page.insert_textbox(
        fitz.Rect(MARGIN, 260, r.width - MARGIN, 360),
        title,
        fontsize=28,
        fontname="helv",
        color=TEXT,
        align=fitz.TEXT_ALIGN_CENTER,
    )
    page.insert_textbox(
        fitz.Rect(MARGIN, 370, r.width - MARGIN, 420),
        "Clube Caixa Secreta · Descubra seu desejo oculto",
        fontsize=12,
        fontname="helv",
        color=TEXT_MUTED,
        align=fitz.TEXT_ALIGN_CENTER,
    )


def draw_cover(doc: fitz.Document) -> None:
    page = doc.new_page(width=595, height=842)
    fill_background(page)
    r = page.rect
    lp = logo_path()
    if lp:
        page.insert_image(fitz.Rect(r.width / 2 - 56, 180, r.width / 2 + 56, 292), filename=str(lp))

    page.insert_textbox(
        fitz.Rect(MARGIN, 320, r.width - MARGIN, 400),
        "Catálogo",
        fontsize=32,
        fontname="helv",
        color=TEXT,
        align=fitz.TEXT_ALIGN_CENTER,
    )
    page.insert_textbox(
        fitz.Rect(MARGIN, 400, r.width - MARGIN, 460),
        "Clube Caixa Secreta",
        fontsize=20,
        fontname="helv",
        color=ACCENT,
        align=fitz.TEXT_ALIGN_CENTER,
    )
    cats = " · ".join(SHEETS_ORDER)
    page.insert_textbox(
        fitz.Rect(MARGIN, 500, r.width - MARGIN, 580),
        cats,
        fontsize=11,
        fontname="helv",
        color=TEXT_MUTED,
        align=fitz.TEXT_ALIGN_CENTER,
    )


def image_rect(inner: fitz.Rect) -> fitz.Rect:
    h = inner.height * 0.48
    return fitz.Rect(inner.x0 + 6, inner.y0 + 6, inner.x1 - 6, inner.y0 + 6 + h)


def draw_product_card(page: fitz.Page, rect: fitz.Rect, product: Product, tmpdir: Path) -> None:
    page.draw_rect(rect, color=CARD_BORDER, width=0.8)
    inner = fitz.Rect(rect.x0 + 4, rect.y0 + 4, rect.x1 - 4, rect.y1 - 4)
    img_box = image_rect(inner)

    if product.image_bytes:
        ext = ".jpg"
        if product.image_bytes[:8] == b"\x89PNG\r\n\x1a\n":
            ext = ".png"
        path = tmpdir / f"img_{id(product)}{ext}"
        path.write_bytes(product.image_bytes)
        try:
            page.insert_image(img_box, filename=str(path), keep_proportion=True)
        except Exception:
            page.draw_rect(img_box, color=TEXT_MUTED, fill=(0.15, 0.1, 0.22))
    else:
        page.draw_rect(img_box, color=TEXT_MUTED, fill=(0.15, 0.1, 0.22))
        page.insert_textbox(
            img_box,
            "Sem foto",
            fontsize=9,
            fontname="helv",
            color=TEXT_MUTED,
            align=fitz.TEXT_ALIGN_CENTER,
        )

    y = img_box.y1 + 8
    name_box = fitz.Rect(inner.x0 + 4, y, inner.x1 - 4, y + 34)
    page.insert_textbox(
        name_box,
        product.name,
        fontsize=10,
        fontname="hebo",
        color=TEXT,
    )

    y = name_box.y1 + 2
    if product.description:
        desc_box = fitz.Rect(inner.x0 + 4, y, inner.x1 - 4, y + 36)
        page.insert_textbox(
            desc_box,
            product.description,
            fontsize=8,
            fontname="helv",
            color=TEXT_MUTED,
        )
        y = desc_box.y1 + 4
    else:
        y += 4

    price_box = fitz.Rect(inner.x0 + 4, y, inner.x1 - 4, y + 22)
    page.insert_textbox(
        price_box,
        f"Meu Preço Venda ({product.price})",
        fontsize=10,
        fontname="hebo",
        color=ACCENT,
    )


def grid_rects(page: fitz.Page) -> list[fitz.Rect]:
    r = page.rect
    usable_w = r.width - 2 * MARGIN
    usable_h = r.height - 2 * MARGIN - 28
    cell_w = (usable_w - GAP) / COLS_PER_ROW
    cell_h = (usable_h - GAP) / ROWS_PER_PAGE
    rects: list[fitz.Rect] = []
    for row in range(ROWS_PER_PAGE):
        for col in range(COLS_PER_ROW):
            x0 = MARGIN + col * (cell_w + GAP)
            y0 = MARGIN + 28 + row * (cell_h + GAP)
            rects.append(fitz.Rect(x0, y0, x0 + cell_w, y0 + cell_h))
    return rects


def draw_product_pages(
    doc: fitz.Document,
    category: str,
    products: list[Product],
    tmpdir: Path,
) -> None:
    for i in range(0, len(products), PRODUCTS_PER_PAGE):
        chunk = products[i : i + PRODUCTS_PER_PAGE]
        page = doc.new_page(width=595, height=842)
        fill_background(page)
        r = page.rect
        page.insert_textbox(
            fitz.Rect(MARGIN, MARGIN, r.width - MARGIN, MARGIN + 22),
            category,
            fontsize=14,
            fontname="hebo",
            color=ACCENT,
        )
        rects = grid_rects(page)
        for rect, product in zip(rects, chunk):
            draw_product_card(page, rect, product, tmpdir)


def resolve_sheets(wb: openpyxl.Workbook, only: list[str] | None) -> list[str]:
    available = {sheet_key(n): n for n in wb.sheetnames}
    if only:
        names = []
        for slug in only:
            key = sheet_key(slug)
            if key in available:
                names.append(available[key])
            else:
                print(f"Aviso: aba não encontrada: {slug}")
        return names
    return [available[sheet_key(s)] for s in SHEETS_ORDER if sheet_key(s) in available]


def generate_pdf(xlsx: Path, output: Path, sheets_filter: list[str] | None = None) -> tuple[Path, int, int]:
    if not xlsx.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {xlsx}")

    site = load_site_catalog_index()
    wb_values = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    wb_images = openpyxl.load_workbook(xlsx, data_only=False)
    sheet_names = resolve_sheets(wb_images, sheets_filter)

    doc = fitz.open()
    draw_cover(doc)
    total = 0
    stats = {"site": 0, "planilha": 0, "sem": 0}

    with tempfile.TemporaryDirectory(prefix="catalogo_pdf_") as tmp:
        tmpdir = Path(tmp)
        for sheet_name in sheet_names:
            ws_values = wb_values[sheet_name]
            ws_images = wb_images[sheet_name]
            products = read_products(ws_values, ws_images, site)
            if not products:
                continue
            for p in products:
                if p.image_source.startswith("site:"):
                    stats["site"] += 1
                elif p.image_source == "planilha":
                    stats["planilha"] += 1
                else:
                    stats["sem"] += 1
            draw_category_page(doc, sheet_name.title())
            draw_product_pages(doc, sheet_name.title(), products, tmpdir)
            total += len(products)
            print(f"  {sheet_name}: {len(products)} produtos")

    wb_values.close()
    wb_images.close()
    output.parent.mkdir(parents=True, exist_ok=True)
    doc.save(str(output))
    page_count = fitz.open(output).page_count
    doc.close()
    print(
        f"  Imagens: {stats['site']} do site · {stats['planilha']} da planilha · "
        f"{stats['sem']} sem foto"
    )
    return output, total, page_count


def main() -> None:
    ap = argparse.ArgumentParser(description="Gera PDF do catálogo a partir da planilha Excel.")
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"Planilha de origem (padrão: {DEFAULT_XLSX})",
    )
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUT,
        help=f"PDF de saída (padrão: {DEFAULT_OUT})",
    )
    ap.add_argument(
        "--abas",
        nargs="*",
        help="Abas a incluir (ex.: comestiveis cosmeticos vibradores velas proteses)",
    )
    args = ap.parse_args()

    print(f"Planilha (textos e preços): {args.xlsx}")
    print("Imagens: site (quando existir) + planilha como reserva\n")
    out, total, pages = generate_pdf(args.xlsx, args.output, args.abas)
    print(f"\nPDF gerado: {out}")
    print(f"Total: {total} produtos em {pages} páginas")


if __name__ == "__main__":
    main()
