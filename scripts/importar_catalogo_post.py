# -*- coding: utf-8 -*-
"""
Importa catálogo a partir da planilha Post + imagens locais.

Planilha: Catalogo_ClubeCaixaSecreta -.xlsx
Imagens: imagens_caixa_secreta/

Categorias no site:
  COMESTÍVEIS  → comestiveis
  COSMÉTICOS + VELAS → cosmeticos
  VIBRADORES + PRÓTESES → vibradores

Uso:
  python scripts/importar_catalogo_post.py
  python scripts/importar_catalogo_post.py --sem-rembg
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import unicodedata
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from PIL import Image

from imagem_fundo import save_transparent_catalog_png, strip_white_edges

BASE = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"C:\Users\Ikaro\post\Catalogo_ClubeCaixaSecreta -.xlsx")
DEFAULT_IMAGES = Path(r"C:\Users\Ikaro\post\imagens_caixa_secreta")
LINGerie_JSON = BASE / "frontend" / "src" / "data" / "catalogo-lingerie.json"
IMPORTADO_JSON = BASE / "frontend" / "src" / "data" / "catalogo-importado.json"
IMG_OUT = BASE / "frontend" / "public" / "importados" / "catalogo"
LOG_FILE = BASE / "dados" / "import_post_sem_imagem.log"

CANVAS = 900
CONTENT_FILL = 0.82

SHEET_TO_SLUG = {
    "COMESTÍVEIS": "comestiveis",
    "COMESTIVEIS": "comestiveis",
    "COSMÉTICOS": "cosmeticos",
    "COSMETICOS": "cosmeticos",
    "VELAS": "cosmeticos",
    "VIBRADORES": "vibradores",
    "PRÓTESES": "vibradores",
    "PROTESES": "vibradores",
}

SHEET_TO_FOLDERS: dict[str, list[str]] = {
    "COMESTÍVEIS": ["COMESTÍVEIS", "COMESTIVEIS"],
    "COMESTIVEIS": ["COMESTÍVEIS", "COMESTIVEIS"],
    "COSMÉTICOS": ["COSMÉTICOS", "COSMETICOS"],
    "COSMETICOS": ["COSMÉTICOS", "COSMETICOS"],
    "VELAS": ["VELAS"],
    "VIBRADORES": ["VIBRADORES"],
    "PRÓTESES": ["PRÓTESES", "PROTESES"],
    "PROTESES": ["PRÓTESES", "PROTESES"],
}

CATEGORY_TITLE = {
    "comestiveis": "Comestíveis",
    "cosmeticos": "Cosméticos",
    "vibradores": "Vibradores",
}

REPLACED_SLUGS = frozenset({"comestiveis", "cosmeticos", "vibradores"})
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def norm_key(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def norm_ref(ref) -> str:
    if ref is None:
        return ""
    return str(ref).strip().upper()


def slugify(text: str, max_len: int = 48) -> str:
    s = unicodedata.normalize("NFKD", text or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return (s or "item")[:max_len]


def parse_price(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def header_index(header: tuple) -> dict[str, int]:
    out = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        key = norm_key(str(h))
        if "produto" in key and "nome" not in out:
            out["nome"] = i
        elif "descricao" in key:
            out["descricao"] = i
        elif key == "foto":
            out["foto"] = i
        elif key == "ref":
            out["ref"] = i
        elif "tamanho" in key:
            out["tamanho"] = i
        elif "sabor" in key or "cor" in key:
            out["sabores"] = i
        elif "preco" in key and "venda" in key:
            out["preco"] = i
        elif "alimentacao" in key:
            out["alimentacao"] = i
    return out


def cell(row: tuple, idx: dict, key: str):
    if key not in idx or idx[key] >= len(row):
        return None
    v = row[idx[key]]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def build_description(desc, tamanho, sabores, alimentacao) -> str:
    parts = []
    if desc:
        parts.append(desc.strip())
    if tamanho and norm_key(tamanho) not in ("unico", ""):
        parts.append(f"Tamanho: {tamanho}")
    if sabores:
        parts.append(str(sabores).strip())
    if alimentacao:
        parts.append(f"Alimentação: {alimentacao}")
    return "\n\n".join(parts)[:2500]


def list_image_files(images_root: Path) -> list[Path]:
    files = []
    for p in images_root.rglob("*"):
        if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
            files.append(p)
    return files


def ref_matches_filename(ref: str, filename: str) -> bool:
    if not ref:
        return False
    ref = ref.strip().upper()
    name = Path(filename).stem.upper()
    if re.match(rf"^{re.escape(ref)}\s*[-–—]", name):
        return True
    if name.startswith(ref + " "):
        return True
    return False


def name_score(product_name: str, filename: str) -> float:
    name = norm_key(Path(filename).stem)
    prod = norm_key(product_name)
    if not prod or not name:
        return 0.0
    if prod in name or name in prod:
        return 0.95
    tokens = [t for t in re.split(r"[^a-z0-9]+", product_name.lower()) if len(t) >= 4]
    if not tokens:
        return 0.0
    hits = sum(1 for t in tokens if t in name)
    return hits / len(tokens)


def find_image(
    product_name: str,
    ref: str | None,
    foto_hint: str | None,
    folder_names: list[str],
    all_files: list[Path],
    images_root: Path,
) -> Path | None:
    pool: list[Path] = []
    for fn in folder_names:
        for variant in (fn, fn.upper(), fn.lower()):
            d = images_root / variant
            if d.is_dir():
                pool.extend(p for p in d.iterdir() if p.suffix.lower() in IMAGE_EXTS)
    if not pool:
        pool = all_files

    if foto_hint:
        hint = Path(foto_hint).name.lower()
        for p in pool:
            if hint in p.name.lower():
                return p

    ref_n = norm_ref(ref)
    if ref_n:
        for p in pool:
            if ref_matches_filename(ref_n, p.name):
                return p

    scored = [(name_score(product_name, p.name), p) for p in pool]
    scored.sort(key=lambda x: -x[0])
    if scored and scored[0][0] >= 0.55:
        return scored[0][1]
    return None


def ensure_rembg():
    try:
        import rembg  # noqa: F401
        return
    except Exception:
        subprocess.check_call(
            [os.environ.get("PYTHON", "python"), "-m", "pip", "install", "-q", "rembg", "onnxruntime", "pillow"]
        )


def remove_background(im: Image.Image, session) -> Image.Image:
    from rembg import remove

    return remove(im.convert("RGBA"), session=session)


def process_image(src: Path, dest: Path, *, use_rembg: bool, session) -> bool:
    try:
        im = Image.open(src)
    except OSError:
        return False
    if use_rembg and session is not None:
        try:
            im = remove_background(im, session)
        except Exception:
            im = im.convert("RGBA")
    else:
        im = im.convert("RGBA")
    im = strip_white_edges(im)
    return save_transparent_catalog_png(im, dest)


def read_products(xlsx: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    products = []
    for sheet_name in wb.sheetnames:
        slug = None
        for key, val in SHEET_TO_SLUG.items():
            if norm_key(sheet_name) == norm_key(key):
                slug = val
                break
        if not slug:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        idx = header_index(header)
        if "nome" not in idx or "preco" not in idx:
            continue
        folders = SHEET_TO_FOLDERS.get(sheet_name, [sheet_name])
        for raw in rows[1:]:
            if not raw or all(c is None or str(c).strip() == "" for c in raw):
                continue
            nome = cell(raw, idx, "nome")
            if not nome:
                continue
            products.append(
                {
                    "sheet": sheet_name,
                    "slug": slug,
                    "name": nome,
                    "description": build_description(
                        cell(raw, idx, "descricao"),
                        cell(raw, idx, "tamanho"),
                        cell(raw, idx, "sabores"),
                        cell(raw, idx, "alimentacao"),
                    ),
                    "ref": cell(raw, idx, "ref"),
                    "price": parse_price(raw[idx["preco"]] if "preco" in idx else 0),
                    "foto_hint": cell(raw, idx, "foto"),
                    "folders": folders,
                }
            )
    wb.close()
    return products


def product_id(slug: str, ref: str | None, name: str) -> str:
    base = slugify(norm_ref(ref) or name)
    return f"imp-{slug}-{base}"[:80]


def load_json(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: list[dict]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def index_by_ref(items: list[dict], slug: str) -> dict[str, dict]:
    out = {}
    for p in items:
        if p.get("categorySlug") != slug:
            continue
        r = norm_ref(p.get("ref"))
        if r:
            out[r] = p
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--imagens", type=Path, default=DEFAULT_IMAGES)
    ap.add_argument("--sem-rembg", action="store_true", help="Não remove fundo (mais rápido)")
    args = ap.parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(args.xlsx)
    if not args.imagens.exists():
        raise FileNotFoundError(args.imagens)

    rows = read_products(args.xlsx)
    print(f"Linhas na planilha: {len(rows)}")

    all_images = list_image_files(args.imagens)
    print(f"Imagens encontradas: {len(all_images)}")

    use_rembg = not args.sem_rembg
    session = None
    if use_rembg:
        ensure_rembg()
        from rembg import new_session

        session = new_session("u2net")
        print("Remoção de fundo: ativa (rembg)")

    lingerie = load_json(LINGerie_JSON)
    importado = load_json(IMPORTADO_JSON)

    for p in lingerie:
        if p.get("categorySlug") in REPLACED_SLUGS:
            p["published"] = False
            p["active"] = False

    kept_importado = [p for p in importado if p.get("categorySlug") not in REPLACED_SLUGS]
    refs_by_slug = {slug: index_by_ref(importado, slug) for slug in REPLACED_SLUGS}

    new_products: list[dict] = []
    sem_imagem: list[str] = []
    updated = created = with_img = 0

    for row in rows:
        slug = row["slug"]
        ref = row.get("ref")
        name = row["name"]
        pid = product_id(slug, ref, name)

        existing = refs_by_slug[slug].get(norm_ref(ref)) if ref else None
        if existing:
            pid = existing["id"]
            updated += 1
        else:
            created += 1

        src = find_image(name, ref, row.get("foto_hint"), row["folders"], all_images, args.imagens)
        image_url = None
        if src:
            dest = IMG_OUT / slug / f"{pid.split('-', 2)[-1] if pid.count('-') >= 2 else slugify(name)}.png"
            safe_name = slugify(norm_ref(ref) or name)[:50]
            dest = IMG_OUT / slug / f"{safe_name}.png"
            if process_image(src, dest, use_rembg=use_rembg, session=session):
                image_url = f"/importados/catalogo/{slug}/{dest.name}"
                with_img += 1
            else:
                sem_imagem.append(f"{name} | ref={ref} | arquivo={src.name}")
        else:
            sem_imagem.append(f"{name} | ref={ref} | SEM ARQUIVO")

        now = utc_now()
        product = {
            "id": pid,
            "name": name,
            "ref": ref,
            "price": row["price"],
            "category": CATEGORY_TITLE[slug],
            "categorySlug": slug,
            "description": row["description"],
            "image": image_url,
            "page": None,
            "source": "Catálogo Post — planilha Clube Caixa Secreta",
            "importSource": "importado",
            "published": bool(image_url),
            "active": True,
            "createdAt": (existing or {}).get("createdAt") or now,
            "updatedAt": now,
        }
        new_products.append(product)

    final_importado = kept_importado + new_products
    save_json(IMPORTADO_JSON, final_importado)
    save_json(LINGerie_JSON, lingerie)

    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    LOG_FILE.write_text("\n".join(sem_imagem) if sem_imagem else "(nenhum)\n", encoding="utf-8")

    print(f"Novos: {created} | Atualizados (por ref): {updated}")
    print(f"Com imagem PNG: {with_img} | Sem imagem: {len(sem_imagem)}")
    print(f"Importado total: {len(final_importado)} (antes removidos slugs: {len(importado) - len(kept_importado)})")
    print(f"Log sem imagem: {LOG_FILE}")


if __name__ == "__main__":
    main()
