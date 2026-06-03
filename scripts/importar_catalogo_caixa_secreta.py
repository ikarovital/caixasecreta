# -*- coding: utf-8 -*-
"""
Importa todo o catálogo de catalogo_caixa_secreta.xlsx + imagens_caixa_secreta.

Uso:
  python scripts/importar_catalogo_caixa_secreta.py
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from PIL import Image

from imagem_fundo import save_transparent_catalog_png

BASE = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"C:\Users\Ikaro\post\catalogo_caixa_secreta.xlsx")
DEFAULT_IMAGES = Path(r"C:\Users\Ikaro\post\imagens_caixa_secreta")
LINGerie_JSON = BASE / "frontend" / "src" / "data" / "catalogo-lingerie.json"
IMPORTADO_JSON = BASE / "frontend" / "src" / "data" / "catalogo-importado.json"
IMG_BASE = BASE / "frontend" / "public" / "importados" / "catalogo"
LOG_FILE = BASE / "dados" / "import_catalogo_caixa_secreta.log"
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
MAX_DESC_LEN = 280

# keys = norm_key da aba da planilha
SHEET_RULES: list[dict] = [
    {"keys": ["comestiveis"], "slug": "comestiveis", "title": "Comestíveis", "folders": ["COMESTÍVEIS", "COMESTIVEIS"], "target": "importado"},
    {"keys": ["cosmeticos"], "slug": "cosmeticos", "title": "Cosméticos", "folders": ["COSMÉTICOS", "COSMETICOS"], "target": "importado"},
    {"keys": ["velas"], "slug": "cosmeticos", "title": "Cosméticos", "folders": ["VELAS"], "target": "importado"},
    {"keys": ["vibradores"], "slug": "vibradores", "title": "Vibradores", "folders": ["VIBRADORES"], "target": "importado"},
    {"keys": ["masturbadoresfem"], "slug": "vibradores", "title": "Vibradores", "folders": ["MASTURBADORES FEM", "MASTURBADORES FEMININO"], "target": "importado"},
    {"keys": ["proteses"], "slug": "proteses", "title": "Próteses", "folders": ["PRÓTESES", "PROTESES"], "target": "importado"},
    {"keys": ["plugs"], "slug": "acessorios", "title": "Acessórios", "folders": ["PLUGS"], "target": "importado"},
    {"keys": ["pompoarismo"], "slug": "acessorios", "title": "Acessórios", "folders": ["POMPOARISMO"], "target": "importado"},
    {"keys": ["acessorios"], "slug": "acessorios", "title": "Acessórios", "folders": ["ACESSÓRIOS", "ACESSORIOS"], "target": "importado"},
    {"keys": ["sadobdsm"], "slug": "sado", "title": "Fetiche e Sado", "folders": ["SADO BDSM", "SADO"], "target": "importado"},
    {"keys": ["calcinha"], "slug": "calcinhas", "title": "Calcinhas", "folders": ["CALCINHA"], "target": "lingerie", "dual": True},
    {"keys": ["lingeries"], "slug": "lingeries", "title": "Lingeries", "folders": ["LINGERIES"], "target": "lingerie", "dual": True},
    {"keys": ["camisolas"], "slug": "camisolas", "title": "Camisolas", "folders": ["CAMISOLAS"], "target": "lingerie"},
    {"keys": ["espartilho"], "slug": "espartilhos", "title": "Espartilhos", "folders": ["ESPARTILHO", "ESPARTILHOS"], "target": "lingerie", "dual": True},
]

IMPORT_SLUGS = frozenset(r["slug"] for r in SHEET_RULES if r["target"] == "importado")
LINGERIE_SLUGS = frozenset(r["slug"] for r in SHEET_RULES if r["target"] == "lingerie")
ALL_SLUGS = IMPORT_SLUGS | LINGERIE_SLUGS


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def norm_key(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def norm_ref(ref) -> str:
    if ref is None:
        return ""
    return str(ref).strip().upper()


def slugify(text: str, max_len: int = 48) -> str:
    s = unicodedata.normalize("NFKD", text or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return (s or "item")[:max_len]


def parse_price(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def header_index(header: tuple) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        key = norm_key(str(h))
        raw_h = str(h).lower()
        if "produto" in key and "nome" not in out:
            out["nome"] = i
        elif "descricao" in key:
            out["descricao"] = i
        elif key == "foto" and "verso" not in raw_h:
            out["foto"] = i
        elif "verso" in key or "fotoverso" in key:
            out["foto_verso"] = i
        elif key == "ref":
            out["ref"] = i
        elif "tamanho" in key:
            out["tamanho"] = i
        elif "sabor" in key or key == "cores" or ("cor" in key and "categoria" not in key):
            out["sabores"] = i
        elif "preco" in key and "venda" in key:
            out["preco"] = i
        elif "alimentacao" in key:
            out["alimentacao"] = i
    return out


def cell(row: tuple, idx: dict, key: str):
    if key not in idx or idx[key] >= len(row):
        return None
    v = row[idx[key]]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_list_field(text: str | None, prefix: str) -> list[str]:
    if not text:
        return []
    s = str(text).strip()
    s = re.sub(rf"^{prefix}\s*:?\s*", "", s, flags=re.I)
    return [p.strip() for p in re.split(r"[,;/]", s) if p.strip()]


def shorten_text(text: str | None, limit: int = MAX_DESC_LEN) -> str:
    if not text:
        return ""
    t = re.sub(r"\s+", " ", str(text).strip())
    if len(t) <= limit:
        return t
    cut = t[:limit].rsplit(" ", 1)[0]
    return (cut or t[:limit]).rstrip(".,; ") + "…"


def build_description(desc, tamanho, sabores, alimentacao) -> str:
    parts = []
    if desc:
        parts.append(shorten_text(desc, 200))
    if tamanho and norm_key(tamanho) not in ("unico", ""):
        parts.append(f"Tamanho: {tamanho.strip()}")
    if sabores:
        parts.append(str(sabores).strip())
    if alimentacao:
        parts.append(f"Alimentação: {alimentacao.strip()}")
    return shorten_text("\n".join(parts), MAX_DESC_LEN)


def rule_for_sheet(sheet_name: str) -> dict | None:
    nk = norm_key(sheet_name)
    for rule in SHEET_RULES:
        if nk in rule["keys"]:
            return rule
    return None


def list_image_files(images_root: Path) -> list[Path]:
    files = []
    for p in images_root.rglob("*"):
        if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
            files.append(p)
    return files


def ref_matches_filename(ref: str, filename: str) -> bool:
    if not ref:
        return False
    ref = ref.strip().upper()
    name = Path(filename).stem.upper()
    if re.match(rf"^{re.escape(ref)}\s*[-–—]", name):
        return True
    if name.startswith(ref + " "):
        return True
    return False


def name_score(product_name: str, filename: str) -> float:
    name = norm_key(Path(filename).stem)
    prod = norm_key(product_name)
    if not prod or not name:
        return 0.0
    if prod in name or name in prod:
        return 0.95
    tokens = [t for t in re.split(r"[^a-z0-9]+", product_name.lower()) if len(t) >= 4]
    if not tokens:
        return 0.0
    return sum(1 for t in tokens if t in name) / len(tokens)


def find_image(
    product_name: str,
    ref: str | None,
    foto_hint: str | None,
    folder_names: list[str],
    all_files: list[Path],
    images_root: Path,
) -> Path | None:
    pool: list[Path] = []
    for fn in folder_names:
        for variant in (fn, fn.upper(), fn.lower()):
            d = images_root / variant
            if d.is_dir():
                pool.extend(p for p in d.iterdir() if p.suffix.lower() in IMAGE_EXTS)
    if not pool:
        pool = all_files

    if foto_hint:
        hint = Path(foto_hint).name.lower()
        for p in pool:
            if hint in p.name.lower():
                return p

    ref_n = norm_ref(ref)
    if ref_n:
        for p in pool:
            if ref_matches_filename(ref_n, p.name):
                return p

    scored = [(name_score(product_name, p.name), p) for p in pool]
    scored.sort(key=lambda x: -x[0])
    if scored and scored[0][0] >= 0.55:
        return scored[0][1]
    return None


def strip_side_label(stem: str) -> str:
    return re.sub(r"\s*-\s*(frente|verso)\s*$", "", stem, flags=re.I).strip()


def list_side_images(folder: Path) -> tuple[list[Path], list[Path]]:
    frente, verso = [], []
    if not folder.is_dir():
        return frente, verso
    for p in folder.iterdir():
        if not p.is_file() or p.suffix.lower() not in IMAGE_EXTS:
            continue
        if "verso" in p.stem.lower():
            verso.append(p)
        elif "frente" in p.stem.lower():
            frente.append(p)
        else:
            frente.append(p)
    return frente, verso


def side_name_score(product_name: str, file_stem: str) -> float:
    prod = norm_key(product_name)
    stem = norm_key(strip_side_label(file_stem))
    if not prod or not stem:
        return 0.0
    if prod == stem or prod in stem or stem in prod:
        return 0.98
    tokens = [t for t in re.split(r"[^a-z0-9]+", product_name.lower()) if len(t) >= 4]
    if not tokens:
        return 0.0
    return sum(1 for t in tokens if t in stem) / len(tokens)


def find_side(product_name: str, ref: str | None, foto_hint: str | None, pool: list[Path]) -> Path | None:
    if foto_hint:
        hint = Path(foto_hint).name.lower()
        for p in pool:
            if hint in p.name.lower():
                return p
    ref_n = norm_ref(ref)
    if ref_n:
        for p in pool:
            if ref_n in p.stem.upper():
                return p
    scored = [(side_name_score(product_name, p.stem), p) for p in pool]
    scored.sort(key=lambda x: -x[0])
    if scored and scored[0][0] >= 0.55:
        return scored[0][1]
    return None


def process_image(src: Path, dest: Path, *, raw_copy: bool) -> bool:
    try:
        im = Image.open(src)
    except OSError:
        return False
    if raw_copy:
        dest.parent.mkdir(parents=True, exist_ok=True)
        if src.suffix.lower() == ".png":
            shutil.copy2(src, dest)
            return dest.exists()
        out = dest if dest.suffix.lower() == ".png" else dest.with_suffix(".png")
        im.convert("RGB").save(out, format="PNG", optimize=True)
        return out.exists()
    return save_transparent_catalog_png(im, dest)


def product_id(slug: str, ref: str | None, name: str) -> str:
    prefix = "imp" if slug in IMPORT_SLUGS else slug
    base = slugify(norm_ref(ref) or name)
    return f"{prefix}-{slug}-{base}"[:80] if prefix == "imp" else f"{slug}-{base}"[:80]


def load_json(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: list[dict]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def read_sheet_rows(ws) -> tuple[dict[str, int], list[tuple]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, []
    return header_index(rows[0]), rows[1:]


def import_products(
    *,
    xlsx_name: str,
    rule: dict,
    sheet_name: str,
    rows: list[tuple],
    idx: dict[str, int],
    images_root: Path,
    all_files: list[Path],
    raw_copy: bool,
    log: list[str],
) -> list[dict]:
    slug = rule["slug"]
    title = rule["title"]
    dual = rule.get("dual", False)
    img_out = IMG_BASE / slug
    img_out.mkdir(parents=True, exist_ok=True)

    frente_pool, verso_pool = [], []
    if dual:
        for fn in rule["folders"]:
            f, v = list_side_images(images_root / fn)
            frente_pool.extend(f)
            verso_pool.extend(v)

    products = []
    for raw in rows:
        if not raw or all(c is None or str(c).strip() == "" for c in raw):
            continue
        nome = cell(raw, idx, "nome")
        if not nome:
            continue
        if "preco" not in idx:
            continue

        ref = cell(raw, idx, "ref")
        tamanho = cell(raw, idx, "tamanho")
        sabores = cell(raw, idx, "sabores")
        alimentacao = cell(raw, idx, "alimentacao")
        desc_raw = cell(raw, idx, "descricao") if "descricao" in idx else None

        row = {
            "name": nome,
            "ref": ref,
            "price": parse_price(raw[idx["preco"]]),
            "description": build_description(desc_raw, tamanho, sabores, alimentacao),
            "size": tamanho,
            "colors": sabores,
            "sizes": parse_list_field(tamanho, "tam"),
            "colorsList": parse_list_field(sabores, "cores"),
            "foto": cell(raw, idx, "foto"),
            "foto_verso": cell(raw, idx, "foto_verso") if "foto_verso" in idx else None,
        }

        safe = slugify(norm_ref(ref) or nome)[:50]
        image_url = image_back = None

        if dual:
            src_frente = find_side(nome, ref, row["foto"], frente_pool)
            src_verso = find_side(nome, ref, row["foto_verso"], verso_pool)
            if src_frente:
                dest = img_out / f"{safe}-frente.png"
                if process_image(src_frente, dest, raw_copy=raw_copy):
                    image_url = f"/importados/catalogo/{slug}/{dest.name}"
                else:
                    log.append(f"[{sheet_name}] {nome} | frente falhou")
            else:
                log.append(f"[{sheet_name}] {nome} | SEM FRENTE")
            if src_verso:
                dest_v = img_out / f"{safe}-verso.png"
                if process_image(src_verso, dest_v, raw_copy=raw_copy):
                    image_back = f"/importados/catalogo/{slug}/{dest_v.name}"
        else:
            src = find_image(nome, ref, row["foto"], rule["folders"], all_files, images_root)
            if src:
                dest = img_out / f"{safe}.png"
                if process_image(src, dest, raw_copy=raw_copy):
                    image_url = f"/importados/catalogo/{slug}/{dest.name}"
                else:
                    log.append(f"[{sheet_name}] {nome} | imagem falhou")
            else:
                log.append(f"[{sheet_name}] {nome} | ref={ref} | SEM IMAGEM")

        now = utc_now()
        pid = product_id(slug, ref, nome)
        products.append(
            {
                "id": pid,
                "name": nome,
                "ref": ref,
                "price": row["price"],
                "category": title,
                "categorySlug": slug,
                "description": row["description"],
                "size": row.get("size"),
                "colors": row.get("colors"),
                "sizes": row.get("sizes") or [],
                "colorsList": row.get("colorsList") or [],
                "image": image_url,
                "imageBack": image_back,
                "page": None,
                "source": f"catalogo_caixa_secreta.xlsx — {sheet_name}",
                "importSource": "planilha-caixa-secreta",
                "published": bool(image_url),
                "active": True,
                "createdAt": now,
                "updatedAt": now,
            }
        )

    pub = sum(1 for p in products if p.get("published"))
    print(f"  {sheet_name} -> {slug}: {len(products)} itens, {pub} com foto")
    return products


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--imagens", type=Path, default=DEFAULT_IMAGES)
    ap.add_argument("--sem-processar-imagem", action="store_true")
    args = ap.parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(args.xlsx)

    from openpyxl import load_workbook

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    all_files = list_image_files(args.imagens)
    print(f"Imagens no disco: {len(all_files)}")

    log: list[str] = []
    novos_importado: list[dict] = []
    novos_lingerie: list[dict] = []

    for sheet_name in wb.sheetnames:
        rule = rule_for_sheet(sheet_name)
        if not rule:
            continue
        ws = wb[sheet_name]
        idx, data_rows = read_sheet_rows(ws)
        if "nome" not in idx or "preco" not in idx:
            print(f"  {sheet_name}: pulada (sem Produto/Preço)")
            continue
        batch = import_products(
            xlsx_name=args.xlsx.name,
            rule=rule,
            sheet_name=sheet_name,
            rows=data_rows,
            idx=idx,
            images_root=args.imagens,
            all_files=all_files,
            raw_copy=args.sem_processar_imagem,
            log=log,
        )
        if rule["target"] == "importado":
            novos_importado.extend(batch)
        else:
            novos_lingerie.extend(batch)

    wb.close()

    lingerie = load_json(LINGerie_JSON)
    importado = load_json(IMPORTADO_JSON)

    lingerie_keep = [p for p in lingerie if p.get("categorySlug") not in ALL_SLUGS]
    importado_keep = [p for p in importado if p.get("categorySlug") not in IMPORT_SLUGS]

    rem_l = len(lingerie) - len(lingerie_keep)
    rem_i = len(importado) - len(importado_keep)
    print(f"Removidos lingerie: {rem_l} | importado: {rem_i}")

    save_json(LINGerie_JSON, lingerie_keep + novos_lingerie)
    save_json(IMPORTADO_JSON, importado_keep + novos_importado)

    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    LOG_FILE.write_text("\n".join(log) if log else "(ok)\n", encoding="utf-8")

    pub_i = sum(1 for p in novos_importado if p.get("published"))
    pub_l = sum(1 for p in novos_lingerie if p.get("published"))
    print(f"Importado: {len(novos_importado)} ({pub_i} com foto)")
    print(f"Lingerie: {len(novos_lingerie)} ({pub_l} com foto)")
    print(f"Total site: {len(lingerie_keep)+len(novos_lingerie)+len(importado_keep)+len(novos_importado)}")
    print(f"Log: {LOG_FILE}")


if __name__ == "__main__":
    main()
