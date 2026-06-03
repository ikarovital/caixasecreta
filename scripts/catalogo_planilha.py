# -*- coding: utf-8 -*-
"""
Utilitários compartilhados: exportar/importar catálogo via planilha Excel.

Colunas da planilha (fonte de controle):
  id, nome, categoria, subcategoria, preco, precoPromocional, estoque,
  published, active, imagem, descricao, origem, importRank, createdAt, updatedAt
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

BASE = Path(__file__).resolve().parents[1]
LINGerie_JSON = BASE / "frontend" / "src" / "data" / "catalogo-lingerie.json"
IMPORTADO_JSON = BASE / "frontend" / "src" / "data" / "catalogo-importado.json"
DEFAULT_XLSX = BASE / "dados" / "catalogo_produtos.xlsx"

SHEET_NAME = "produtos"

COLUMNS = [
    "id",
    "nome",
    "categoria",
    "subcategoria",
    "preco",
    "precoPromocional",
    "estoque",
    "published",
    "active",
    "imagem",
    "descricao",
    "origem",
    "importRank",
    "createdAt",
    "updatedAt",
]

CATEGORY_SLUG_TO_TITLE = {
    "calcinhas": "Calcinhas",
    "conjuntos": "Conjuntos",
    "espartilhos": "Espartilhos",
    "fantasias": "Fantasias",
    "comestiveis": "Comestíveis",
    "cosmeticos": "Cosméticos",
    "vibradores": "Vibradores",
    "acessorios": "Acessórios",
    "sex-shop": "Sex Shop",
    "sado": "Fetiche e Sado",
}


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def load_json_list(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def save_json_list(path: Path, items: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")


def catalog_file_for_product(product_id: str) -> Path:
    if product_id.startswith("imp-") or product_id.startswith("miess-"):
        return IMPORTADO_JSON
    return LINGerie_JSON


def load_catalog_index() -> dict[str, tuple[dict, Path]]:
    """id -> (product dict ref, arquivo origem)"""
    index: dict[str, tuple[dict, Path]] = {}
    for path in (LINGerie_JSON, IMPORTADO_JSON):
        for p in load_json_list(path):
            pid = p.get("id")
            if pid:
                index[pid] = (p, path)
    return index


def load_all_products() -> list[dict]:
    return load_json_list(LINGerie_JSON) + load_json_list(IMPORTADO_JSON)


def product_origin(product: dict) -> str:
    if product.get("importSource") in ("importado", "miess"):
        return "importado"
    src = product.get("source") or ""
    return src if src else "catalogo"


def product_to_row(product: dict) -> dict[str, Any]:
    slug = product.get("categorySlug") or ""
    category_title = product.get("category") or CATEGORY_SLUG_TO_TITLE.get(slug, slug)
    return {
        "id": product.get("id") or "",
        "nome": product.get("name") or "",
        "categoria": category_title,
        "subcategoria": slug,
        "preco": product.get("price"),
        "precoPromocional": product.get("listPrice"),
        "estoque": product.get("stock"),
        "published": product.get("published", False),
        "active": product.get("active", True),
        "imagem": product.get("image") or "",
        "descricao": product.get("description") or "",
        "origem": product_origin(product),
        "importRank": product.get("importRank"),
        "createdAt": product.get("createdAt") or "",
        "updatedAt": product.get("updatedAt") or "",
    }


def parse_bool(value: Any, default: bool | None = None) -> bool | None:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(int(value))
    s = str(value).strip().lower()
    if s in ("true", "verdadeiro", "sim", "s", "1", "yes", "y"):
        return True
    if s in ("false", "falso", "nao", "não", "n", "0", "no"):
        return False
    return default


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("R$", "").replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def parse_datetime_cell(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).replace(microsecond=0).isoformat()
    s = str(value).strip()
    return s or None


def row_to_product_updates(row: dict[str, Any]) -> dict[str, Any]:
    """Campos internos a partir de uma linha da planilha."""
    slug = (row.get("subcategoria") or "").strip()
    category_title = (row.get("categoria") or "").strip()
    if slug and not category_title:
        category_title = CATEGORY_SLUG_TO_TITLE.get(slug, slug.replace("-", " ").title())
    if category_title and not slug:
        slug = category_title.lower().replace(" ", "-")

    updates: dict[str, Any] = {
        "name": (row.get("nome") or "").strip(),
        "category": category_title or None,
        "categorySlug": slug or None,
        "description": (row.get("descricao") or "").strip(),
        "image": (row.get("imagem") or "").strip() or None,
        "source": (row.get("origem") or "").strip() or None,
    }

    price = parse_number(row.get("preco"))
    if price is not None:
        updates["price"] = price

    list_price = parse_number(row.get("precoPromocional"))
    if list_price is not None:
        updates["listPrice"] = list_price
    elif row.get("precoPromocional") == "":
        updates["listPrice"] = None

    stock = parse_int(row.get("estoque"))
    if stock is not None:
        updates["stock"] = stock
    elif row.get("estoque") == "":
        updates["stock"] = None

    published = parse_bool(row.get("published"))
    if published is not None:
        updates["published"] = published

    active = parse_bool(row.get("active"))
    if active is not None:
        updates["active"] = active

    rank = parse_int(row.get("importRank"))
    if rank is not None:
        updates["importRank"] = rank
    elif row.get("importRank") == "":
        updates["importRank"] = None

    created = parse_datetime_cell(row.get("createdAt"))
    if created:
        updates["createdAt"] = created

    updated = parse_datetime_cell(row.get("updatedAt"))
    if updated:
        updates["updatedAt"] = updated

    return {k: v for k, v in updates.items() if v is not None or k in ("listPrice", "stock", "importRank", "image")}


def ensure_catalog_defaults(product: dict) -> dict:
    if "published" not in product:
        product["published"] = bool(product.get("image"))
    if "active" not in product:
        product["active"] = True
    if "createdAt" not in product or not product["createdAt"]:
        product["createdAt"] = utc_now_iso()
    if "updatedAt" not in product:
        product["updatedAt"] = product["createdAt"]
    return product


def merge_product(existing: dict | None, updates: dict[str, Any], *, touch_updated: bool = True) -> dict:
    base = dict(existing or {})
    base.update(updates)
    ensure_catalog_defaults(base)
    if touch_updated:
        base["updatedAt"] = utc_now_iso()
    return base


def new_product_from_row(row: dict[str, Any]) -> dict:
    pid = (row.get("id") or "").strip()
    if not pid:
        raise ValueError("Linha sem id — obrigatório para novos produtos.")
    now = utc_now_iso()
    product = merge_product(
        {"id": pid},
        row_to_product_updates(row),
        touch_updated=False,
    )
    product.setdefault("createdAt", parse_datetime_cell(row.get("createdAt")) or now)
    product["updatedAt"] = parse_datetime_cell(row.get("updatedAt")) or now
    return product


def read_sheet_rows(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        wb.close()
        return []

    col_index = {name: i for i, name in enumerate(header) if name}
    missing = [c for c in COLUMNS if c not in col_index]
    if missing:
        wb.close()
        raise ValueError(f"Colunas ausentes na planilha: {', '.join(missing)}")

    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        if not raw or all(c is None or str(c).strip() == "" for c in raw):
            continue
        row = {col: raw[col_index[col]] if col_index[col] < len(raw) else None for col in COLUMNS}
        if not str(row.get("id") or "").strip():
            continue
        out.append(row)
    wb.close()
    return out


def write_sheet(path: Path, products: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for product in sorted(products, key=lambda p: (p.get("categorySlug") or "", p.get("name") or "")):
        row = product_to_row(product)
        ws.append([row[col] for col in COLUMNS])

    widths = {
        "id": 36,
        "nome": 42,
        "categoria": 18,
        "subcategoria": 16,
        "preco": 10,
        "precoPromocional": 14,
        "estoque": 10,
        "published": 10,
        "active": 10,
        "imagem": 48,
        "descricao": 50,
        "origem": 22,
        "importRank": 12,
        "createdAt": 22,
        "updatedAt": 22,
    }
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 14)

    ws.freeze_panes = "A2"
    wb.save(path)


def find_duplicate_ids(products: list[dict]) -> list[str]:
    seen: set[str] = set()
    dups: list[str] = []
    for p in products:
        pid = p.get("id")
        if not pid:
            continue
        if pid in seen:
            dups.append(pid)
        seen.add(pid)
    return dups


def export_catalog(path: Path = DEFAULT_XLSX) -> tuple[int, Path]:
    products = load_all_products()
    dups = find_duplicate_ids(products)
    if dups:
        print(f"AVISO: {len(dups)} id(s) duplicado(s) no JSON (exportados mesmo assim): {dups[:5]}")
    for p in products:
        ensure_catalog_defaults(p)
    write_sheet(path, products)
    return len(products), path


def import_catalog(path: Path = DEFAULT_XLSX) -> dict[str, int]:
    if not path.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {path}")

    rows = read_sheet_rows(path)
    row_ids = [str(r.get("id")).strip() for r in rows]
    dup_rows = find_duplicate_ids([{"id": i} for i in row_ids])
    if dup_rows:
        print(f"AVISO: ids repetidos na planilha (última linha prevalece): {dup_rows[:5]}")

    index = load_catalog_index()

    lingerie = load_json_list(LINGerie_JSON)
    importado = load_json_list(IMPORTADO_JSON)
    by_id_l = {p["id"]: p for p in lingerie if p.get("id")}
    by_id_i = {p["id"]: p for p in importado if p.get("id")}

    updated = created = skipped = 0

    for row in rows:
        pid = str(row.get("id")).strip()
        updates = row_to_product_updates(row)

        if pid in index:
            existing, file_path = index[pid]
            merged = merge_product(existing, updates)
            if file_path == IMPORTADO_JSON:
                by_id_i[pid] = merged
            else:
                by_id_l[pid] = merged
            updated += 1
        else:
            product = new_product_from_row(row)
            target = catalog_file_for_product(pid)
            if target == IMPORTADO_JSON:
                by_id_i[pid] = product
            else:
                by_id_l[pid] = product
            created += 1

    save_json_list(LINGerie_JSON, list(by_id_l.values()))
    save_json_list(IMPORTADO_JSON, list(by_id_i.values()))

    return {
        "rows": len(rows),
        "updated": updated,
        "created": created,
        "skipped": skipped,
        "total_lingerie": len(by_id_l),
        "total_importado": len(by_id_i),
    }
